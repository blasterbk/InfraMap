import os
import csv
import glob
import json
import networkx as nx
from collections import defaultdict

def parse_linode(data_dir, G, servers_by_ip):
    linode_dir = os.path.join(data_dir, 'linode')
    if not os.path.exists(linode_dir):
        return
        
    # Support CSV files
    linode_files = glob.glob(os.path.join(linode_dir, '*.csv'))
    for linodes_file in linode_files:
        with open(linodes_file, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f, delimiter='\t')
            for row in reader:
                _process_linode_row(row, G, servers_by_ip)
                
    # Support Excel files (.xlsx)
    xlsx_files = glob.glob(os.path.join(linode_dir, '*.xlsx'))
    if xlsx_files:
        try:
            import openpyxl
            for xlsx_file in xlsx_files:
                wb = openpyxl.load_workbook(xlsx_file, data_only=True)
                ws = wb.active
                headers = [str(cell.value).strip().lower() if cell.value else f"col_{i}" for i, cell in enumerate(ws[1])]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_dict = dict(zip(headers, [str(v) if v is not None else "" for v in row]))
                    _process_linode_row(row_dict, G, servers_by_ip)
        except ImportError:
            print("Warning: openpyxl is not installed. Please run `pip install openpyxl` to parse Linode.xlsx.")

def _process_linode_row(row, G, servers_by_ip):
    if not row or 'label' not in row or not row['label'].strip():
        return
        
    name = row['label']
    ips_raw = row.get('ipv4', '')
    region = row.get('region', '')
    tag = row.get('tags', '')
    
    image = row.get('image', '')
    status = row.get('status', '')
    instance_type = row.get('type', '')
    vcpus = row.get('vcpus', '')
    
    ips = [ip.strip(' "') for ip in ips_raw.split(',') if ip.strip(' "')]
    ip_str = ", ".join(ips)
    
    source_info = (
        f"IP(s): {ip_str} | Region: {region} | Tags: {tag} | "
        f"Image: {image} | Status: {status} | InstType: {instance_type} | vCPUs: {vcpus}"
    )
    
    G.add_node(name, 
               type="Server", 
               ip=ip_str,
               label=f"{name}\\n({ip_str})",
               file_type="Linode Server",
               source_file=source_info,
               shape='icon',
               icon={'face': '"Font Awesome 6 Free"', 'code': '\uf233', 'weight': '900', 'color': '#4E79A7'},
               provider="linode")
    
    for ip in ips:
        servers_by_ip[ip] = name
    
    if region:
        G.add_node(region, type="Region", label=f"{region}\\n(Region)", file_type="Region", source_file="-",
                   shape='icon', icon={'face': '"Font Awesome 6 Free"', 'code': '\uf3c5', 'weight': '900', 'color': '#F28E2B'}, provider="linode")
        G.add_edge(name, region, relation="deployed_in")
    
    if tag:
        G.add_node(tag, type="Environment", label=f"{tag}\\n(Environment)", file_type="Environment", source_file="-",
                   shape='icon', icon={'face': '"Font Awesome 6 Free"', 'code': '\uf02b', 'weight': '900', 'color': '#E15759'}, provider="linode")
        G.add_edge(name, tag, relation="has_tag")

def parse_bunny(data_dir, G, servers_by_ip):
    bunny_dir = os.path.join(data_dir, 'Bunny')
    if not os.path.exists(bunny_dir):
        return
        
    bind_files = glob.glob(os.path.join(bunny_dir, '*.bind'))
    
    for bind_file in bind_files:
        filename = os.path.basename(bind_file)
        name_parts = filename.split('.')
        if len(name_parts) >= 4 and name_parts[-1] == 'bind' and '-' in name_parts[-2]:
            main_domain = ".".join(name_parts[:-2])
        else:
            main_domain = ".".join(name_parts[:-1]) # Fallback
            
        G.add_node(main_domain, type="MainDomain", label=f"{main_domain}\\n(Main Domain)", file_type="Bunny MainDomain", source_file=f"Zone File: {filename}",
                   shape='icon', icon={'face': '"Font Awesome 6 Free"', 'code': '\uf0ac', 'weight': '900', 'color': '#76B7B2'}, provider=f"bunny_{main_domain}")

        domain_ips = defaultdict(list)
        with open(bind_file, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith(';'):
                    continue
                
                line_parts = line.split()
                if len(line_parts) >= 5 and line_parts[3] == 'A':
                    domain = line_parts[0].rstrip('.')
                    target_ip = line_parts[4]
                    if target_ip not in domain_ips[domain]:
                        domain_ips[domain].append(target_ip)
                        
        for domain, ips in domain_ips.items():
            is_main = (domain == main_domain)
            node_type = "MainDomain" if is_main else "SubDomain"
            
            ip_str = ", ".join(ips)
            label_ip_str = ip_str if len(ips) <= 2 else f"{ips[0]}, ... (+{len(ips)-1})"
            
            has_valid_target = False
            for target_ip in ips:
                if target_ip in servers_by_ip:
                    linode_server = servers_by_ip[target_ip]
                    G.add_edge(domain, linode_server, relation="routes_traffic_to")
                    has_valid_target = True
            
            color = '#76B7B2' if is_main else '#59A14F'
            
            is_orphaned = not has_valid_target
            if is_orphaned:
                color = '#ff4444' # RED!
                label = f"⚠️ {domain}\\n(ORPHANED)"
            else:
                label = f"{domain}\\n({label_ip_str})"
            
            G.add_node(domain, 
                       type=node_type, 
                       ip=ip_str,
                       label=label,
                       color=color, # Override color
                       file_type=f"Bunny {node_type}",
                       source_file=f"Target IPs: {ip_str} | Zone: {main_domain}",
                       shape='icon', icon={'face': '"Font Awesome 6 Free"', 'code': '\uf0ac', 'weight': '900', 'color': color},
                       provider=f"bunny_{main_domain}")
            
            if not is_main:
                G.add_edge(domain, main_domain, relation="subdomain_of")

def parse_kubernetes(data_dir, G, servers_by_ip, cluster_filename=None):
    k8s_dir = os.path.join(data_dir, "kubernetes")
    if not os.path.exists(k8s_dir):
        return
        
    STANDARD_NS_COLORS = {
        "kube-system": "#9E9E9E",
        "default": "#8AC926",
        "monitoring": "#1982C4",
        "production": "#FF595E",
        "staging": "#FFCA3A",
    }
    
    COLORS = ["#6A4C93", "#F15BB5", "#00F5D4", "#9B5DE5", "#FEE440", "#00BBF9"]
    ns_colors = {}
    def get_ns_color(ns):
        if ns in STANDARD_NS_COLORS:
            return STANDARD_NS_COLORS[ns]
        if ns not in ns_colors:
            ns_colors[ns] = COLORS[len(ns_colors) % len(COLORS)]
        return ns_colors[ns]

    CLUSTER_OFFSETS = [
        (-1800, 0),
        (1800, 0),
        (0, 1800),
        (0, -1800),
    ]

    cluster_files = sorted([f for f in os.listdir(k8s_dir) if f.endswith(".json")])
    if cluster_filename:
        cluster_files = [f for f in cluster_files if f == cluster_filename]

    for cluster_index, filename in enumerate(cluster_files):
            
        filepath = os.path.join(k8s_dir, filename)
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)
        except Exception as e:
            print(f"Error parsing {filepath}: {e}")
            continue
            
        items = data.get("items", [])
        if not items:
            continue
            
        nodes = []
        services = []
        ingresses = []
        pods = []
        pvcs = []
        configmaps = []
        secrets = []
        
        deployments = []
        statefulsets = []
        daemonsets = []
        rolebindings = []
        clusterrolebindings = []
        networkpolicies = []
        
        for item in items:
            kind = item.get("kind")
            if kind == "Node":
                nodes.append(item)
            elif kind == "Service":
                services.append(item)
            elif kind == "Ingress":
                ingresses.append(item)
            elif kind == "Pod":
                pods.append(item)
            elif kind == "PersistentVolumeClaim":
                pvcs.append(item)
            elif kind == "ConfigMap":
                configmaps.append(item)
            elif kind == "Secret":
                secrets.append(item)
            elif kind == "Deployment":
                deployments.append(item)
            elif kind == "StatefulSet":
                statefulsets.append(item)
            elif kind == "DaemonSet":
                daemonsets.append(item)
            elif kind == "RoleBinding":
                rolebindings.append(item)
            elif kind == "ClusterRoleBinding":
                clusterrolebindings.append(item)
            elif kind == "NetworkPolicy":
                networkpolicies.append(item)

        cluster_name = filename.replace('.json', '')
        cluster_id = f"k8s_cluster_{cluster_name}"
        cx, cy = CLUSTER_OFFSETS[cluster_index % len(CLUSTER_OFFSETS)]


        def make_node(nid, **kwargs):
            kwargs["cluster_group"] = cluster_index
            kwargs["cluster_x"] = cx + (hash(nid) % 800 - 400)
            kwargs["cluster_y"] = cy + (hash(nid + "y") % 800 - 400)
            G.add_node(nid, **kwargs)

        make_node(cluster_id, type="MainDomain", label=f"K8s Cluster\\n({cluster_name})", file_type="K8s Cluster", source_file="-", shape='icon', icon={'face': '"Font Awesome 6 Free"', 'code': '\uf6ff', 'weight': '900', 'color': '#06D6A0'}, provider="kubernetes", size=25)

        # Physical Infrastructure: Kubernetes Nodes
        for n in nodes:
            n_name = n["metadata"]["name"]
            n_id = f"node_{cluster_name}_{n_name}"
            if not G.has_node(n_id):
                make_node(n_id, type="Server", label=f"{n_name}\\n(Server Node)", file_type="K8s Node", source_file="-", shape='icon', icon={'face': '"Font Awesome 6 Free"', 'code': '\uf233', 'weight': '900', 'color': '#118AB2'}, provider="kubernetes", namespace="system", size=25)

        namespaces = set()

        for s in services:
            name = s["metadata"]["name"]
            namespace = s["metadata"].get("namespace", "default")
            namespaces.add(namespace)
            selector = s.get("spec", {}).get("selector", {})
            
            svc_id = f"svc_{cluster_name}_{namespace}_{name}"
            make_node(svc_id, type="Region", label=f"{name}\\n(Service)", file_type="K8s Service", source_file=f"Namespace: {namespace}", shape='icon', icon={'face': '"Font Awesome 6 Free"', 'code': '\uf085', 'weight': '900', 'color': get_ns_color(namespace)}, provider="kubernetes", namespace=namespace, size=15)
            
            spec = s.get("spec", {})
            if spec.get("type") == "ExternalName":
                ext = spec.get("externalName")
                if ext:
                    parts = ext.split('.')
                    if len(parts) >= 2:
                        target_svc = parts[0]
                        target_ns = parts[1]
                        
                        target_ns_id = f"ns_{cluster_name}_{target_ns}"
                        if not G.has_node(target_ns_id):
                            make_node(target_ns_id, type="SubDomain", label=f"{target_ns}\\n(Namespace)", file_type="K8s Namespace", source_file="-", shape='icon', icon={'face': '"Font Awesome 6 Free"', 'code': '\uf07b', 'weight': '900', 'color': get_ns_color(target_ns)}, provider="kubernetes", namespace=target_ns, size=20)
                            
                        target_svc_id = f"svc_{cluster_name}_{target_ns}_{target_svc}"
                        G.add_edge(svc_id, target_ns_id, relation="bridges_to_ns")
                        G.add_edge(target_ns_id, target_svc_id, relation="contains_svc")
            
            if selector:
                for p in pods:
                    if p["metadata"].get("namespace", "default") != namespace:
                        continue
                    labels = p["metadata"].get("labels", {})
                    match = True
                    for k, v in selector.items():
                        if labels.get(k) != v:
                            match = False
                            break
                    if match:
                        pod_name = p["metadata"]["name"]
                        pod_id = f"pod_{cluster_name}_{namespace}_{pod_name}"
                        if not G.has_node(pod_id):
                            make_node(pod_id, type="Region", label=f"{pod_name}\\n(Pod)", file_type="K8s Pod", source_file="-", shape='icon', icon={'face': '"Font Awesome 6 Free"', 'code': '\uf1b2', 'weight': '900', 'color': get_ns_color(namespace)}, provider="kubernetes", namespace=namespace, size=12)
                        
                        G.add_edge(svc_id, pod_id, relation="selects_pod")
                        
                        containers = p.get("spec", {}).get("containers", [])
                        for c in containers:
                            img = c.get("image")
                            if img:
                                img_id = f"img_{cluster_name}_{img}"
                                if not G.has_node(img_id):
                                    make_node(img_id, type="Environment", label=f"{img}\\n(Docker Image)", file_type="Docker Image", source_file="-", shape='icon', icon={'face': '"Font Awesome 6 Brands"', 'code': '\uf395', 'weight': '400', 'color': get_ns_color(namespace)}, provider="kubernetes", namespace=namespace, size=10)
                                G.add_edge(pod_id, img_id, relation="runs_image")
                                
                        volumes = p.get("spec", {}).get("volumes", [])
                        for v in volumes:
                            if "persistentVolumeClaim" in v:
                                claim_name = v["persistentVolumeClaim"].get("claimName")
                                if claim_name:
                                    pvc_id = f"pvc_{cluster_name}_{namespace}_{claim_name}"
                                    if not G.has_node(pvc_id):
                                        make_node(pvc_id, type="Environment", label=f"{claim_name}\\n(PVC)", file_type="PVC", source_file="-", shape='icon', icon={'face': '"Font Awesome 6 Free"', 'code': '\uf1c0', 'weight': '900', 'color': get_ns_color(namespace)}, provider="kubernetes", namespace=namespace, size=15)
                                    G.add_edge(pod_id, pvc_id, relation="mounts_pvc")
                            if "configMap" in v:
                                cm_name = v["configMap"].get("name")
                                if cm_name:
                                    cm_id = f"cm_{cluster_name}_{namespace}_{cm_name}"
                                    if not G.has_node(cm_id):
                                        make_node(cm_id, type="Environment", label=f"{cm_name}\\n(ConfigMap)", file_type="ConfigMap", source_file="-", shape='icon', icon={'face': '"Font Awesome 6 Free"', 'code': '\uf013', 'weight': '900', 'color': get_ns_color(namespace)}, provider="kubernetes", namespace=namespace, size=10)
                                    G.add_edge(pod_id, cm_id, relation="mounts_cm")
                            if "secret" in v:
                                sec_name = v["secret"].get("secretName")
                                if sec_name:
                                    sec_id = f"sec_{cluster_name}_{namespace}_{sec_name}"
                                    if not G.has_node(sec_id):
                                        make_node(sec_id, type="Environment", label=f"{sec_name}\\n(Secret)", file_type="Secret", source_file="-", shape='icon', icon={'face': '"Font Awesome 6 Free"', 'code': '\uf084', 'weight': '900', 'color': get_ns_color(namespace)}, provider="kubernetes", namespace=namespace, size=10)
                                    G.add_edge(pod_id, sec_id, relation="mounts_secret")
                            if "projected" in v:
                                for src in v.get("projected", {}).get("sources", []):
                                    if "configMap" in src:
                                        cm_name = src["configMap"].get("name")
                                        if cm_name:
                                            cm_id = f"cm_{cluster_name}_{namespace}_{cm_name}"
                                            if not G.has_node(cm_id):
                                                make_node(cm_id, type="Environment", label=f"{cm_name}\\n(ConfigMap)", file_type="ConfigMap", source_file="-", shape='icon', icon={'face': '"Font Awesome 6 Free"', 'code': '\uf013', 'weight': '900', 'color': get_ns_color(namespace)}, provider="kubernetes", namespace=namespace, size=10)
                                            G.add_edge(pod_id, cm_id, relation="mounts_cm")
                                    if "secret" in src:
                                        sec_name = src["secret"].get("name")
                                        if sec_name:
                                            sec_id = f"sec_{cluster_name}_{namespace}_{sec_name}"
                                            if not G.has_node(sec_id):
                                                make_node(sec_id, type="Environment", label=f"{sec_name}\\n(Secret)", file_type="Secret", source_file="-", shape='icon', icon={'face': '"Font Awesome 6 Free"', 'code': '\uf084', 'weight': '900', 'color': get_ns_color(namespace)}, provider="kubernetes", namespace=namespace, size=10)
                                            G.add_edge(pod_id, sec_id, relation="mounts_secret")
                        
                        for c in containers:
                            env_from = c.get("envFrom", [])
                            for ef in env_from:
                                if "configMapRef" in ef:
                                    cm_name = ef["configMapRef"].get("name")
                                    if cm_name:
                                        cm_id = f"cm_{cluster_name}_{namespace}_{cm_name}"
                                        if not G.has_node(cm_id):
                                            make_node(cm_id, type="Environment", label=f"{cm_name}\\n(ConfigMap)", file_type="ConfigMap", source_file="-", shape='icon', icon={'face': '"Font Awesome 6 Free"', 'code': '\uf013', 'weight': '900', 'color': get_ns_color(namespace)}, provider="kubernetes", namespace=namespace, size=10)
                                        G.add_edge(pod_id, cm_id, relation="uses_env_cm")
                                if "secretRef" in ef:
                                    sec_name = ef["secretRef"].get("name")
                                    if sec_name:
                                        sec_id = f"sec_{cluster_name}_{namespace}_{sec_name}"
                                        if not G.has_node(sec_id):
                                            make_node(sec_id, type="Environment", label=f"{sec_name}\\n(Secret)", file_type="Secret", source_file="-", shape='icon', icon={'face': '"Font Awesome 6 Free"', 'code': '\uf084', 'weight': '900', 'color': get_ns_color(namespace)}, provider="kubernetes", namespace=namespace, size=10)
                                        G.add_edge(pod_id, sec_id, relation="uses_env_secret")
                                        
                            env = c.get("env", [])
                            for ev in env:
                                value_from = ev.get("valueFrom", {})
                                if "configMapKeyRef" in value_from:
                                    cm_name = value_from["configMapKeyRef"].get("name")
                                    if cm_name:
                                        cm_id = f"cm_{cluster_name}_{namespace}_{cm_name}"
                                        if not G.has_node(cm_id):
                                            make_node(cm_id, type="Environment", label=f"{cm_name}\\n(ConfigMap)", file_type="ConfigMap", source_file="-", shape='icon', icon={'face': '"Font Awesome 6 Free"', 'code': '\uf013', 'weight': '900', 'color': get_ns_color(namespace)}, provider="kubernetes", namespace=namespace, size=10)
                                        G.add_edge(pod_id, cm_id, relation="uses_env_cm")
                                if "secretKeyRef" in value_from:
                                    sec_name = value_from["secretKeyRef"].get("name")
                                    if sec_name:
                                        sec_id = f"sec_{cluster_name}_{namespace}_{sec_name}"
                                        if not G.has_node(sec_id):
                                            make_node(sec_id, type="Environment", label=f"{sec_name}\\n(Secret)", file_type="Secret", source_file="-", shape='icon', icon={'face': '"Font Awesome 6 Free"', 'code': '\uf084', 'weight': '900', 'color': get_ns_color(namespace)}, provider="kubernetes", namespace=namespace, size=10)
                                        G.add_edge(pod_id, sec_id, relation="uses_env_secret")

        for i in ingresses:
            name = i["metadata"]["name"]
            namespace = i["metadata"].get("namespace", "default")
            namespaces.add(namespace)
            
            rules = i.get("spec", {}).get("rules", [])
            for r in rules:
                host = r.get("host")
                if not host:
                    continue
                
                ns_id = f"ns_{cluster_name}_{namespace}"
                if not G.has_node(ns_id):
                    make_node(ns_id, type="SubDomain", label=f"{namespace}\\n(Namespace)", file_type="K8s Namespace", source_file="-", shape='icon', icon={'face': '"Font Awesome 6 Free"', 'code': '\uf07b', 'weight': '900', 'color': get_ns_color(namespace)}, provider="kubernetes", namespace=namespace, size=20)
                    G.add_edge(cluster_id, ns_id, relation="contains_ns")

                host_id = f"domain_{cluster_name}_{host}"
                if not G.has_node(host_id):
                    make_node(host_id, type="MainDomain", label=f"{host}\\n(Domain)", file_type="Domain", source_file="-", shape='icon', icon={'face': '"Font Awesome 6 Free"', 'code': '\uf0ac', 'weight': '900', 'color': get_ns_color(namespace)}, provider="kubernetes", namespace=namespace, size=15)
                
                G.add_edge(ns_id, host_id, relation="hosts_domain")
                
                paths = r.get("http", {}).get("paths", [])
                for p in paths:
                    backend_svc = p.get("backend", {}).get("service", {}).get("name")
                    if backend_svc:
                        svc_id = f"svc_{cluster_name}_{namespace}_{backend_svc}"
                        if G.has_node(svc_id):
                            G.add_edge(host_id, svc_id, relation="routes_to")

        for c_list, kind_name, icon, color in [
            (deployments, "Deployment", '\uf0e8', "#F15BB5"),
            (statefulsets, "StatefulSet", '\uf1c0', "#FEE440"),
            (daemonsets, "DaemonSet", '\uf233', "#9B5DE5")
        ]:
            for c in c_list:
                name = c["metadata"]["name"]
                namespace = c["metadata"].get("namespace", "default")
                
                ctrl_id = f"{kind_name.lower()}_{cluster_name}_{namespace}_{name}"
                make_node(ctrl_id, type="Region", label=f"{name}\\n({kind_name})", file_type=f"K8s {kind_name}", source_file=f"Namespace: {namespace}", shape='icon', icon={'face': '"Font Awesome 6 Free"', 'code': icon, 'weight': '900', 'color': color}, provider="kubernetes", namespace=namespace, size=18)
                
                selector = c.get("spec", {}).get("selector", {}).get("matchLabels", {})
                if selector:
                    for p in pods:
                        if p["metadata"].get("namespace", "default") != namespace:
                            continue
                        labels = p["metadata"].get("labels", {})
                        match = True
                        for k, v in selector.items():
                            if labels.get(k) != v:
                                match = False
                                break
                        if match:
                            pod_name = p["metadata"]["name"]
                            pod_id = f"pod_{cluster_name}_{namespace}_{pod_name}"
                            is_new = not G.has_node(pod_id)
                            if is_new:
                                make_node(pod_id, type="Region", label=f"{pod_name}\\n(Pod)", file_type="K8s Pod", source_file="-", shape='icon', icon={'face': '"Font Awesome 6 Free"', 'code': '\uf1b2', 'weight': '900', 'color': get_ns_color(namespace)}, provider="kubernetes", namespace=namespace, size=12)
                                
                                containers = p.get("spec", {}).get("containers", [])
                                for ctr in containers:
                                    img = ctr.get("image")
                                    if img:
                                        img_id = f"img_{cluster_name}_{img}"
                                        if not G.has_node(img_id):
                                            make_node(img_id, type="Environment", label=f"{img}\\n(Docker Image)", file_type="Docker Image", source_file="-", shape='icon', icon={'face': '"Font Awesome 6 Brands"', 'code': '\uf395', 'weight': '400', 'color': get_ns_color(namespace)}, provider="kubernetes", namespace=namespace, size=10)
                                        G.add_edge(pod_id, img_id, relation="runs_image")
                                        
                                volumes = p.get("spec", {}).get("volumes", [])
                                for v in volumes:
                                    if "persistentVolumeClaim" in v:
                                        claim_name = v["persistentVolumeClaim"].get("claimName")
                                        if claim_name:
                                            pvc_id = f"pvc_{cluster_name}_{namespace}_{claim_name}"
                                            if not G.has_node(pvc_id):
                                                make_node(pvc_id, type="Environment", label=f"{claim_name}\\n(PVC)", file_type="PVC", source_file="-", shape='icon', icon={'face': '"Font Awesome 6 Free"', 'code': '\uf1c0', 'weight': '900', 'color': get_ns_color(namespace)}, provider="kubernetes", namespace=namespace, size=15)
                                            G.add_edge(pod_id, pvc_id, relation="mounts_pvc")
                                    if "configMap" in v:
                                        cm_name = v["configMap"].get("name")
                                        if cm_name:
                                            cm_id = f"cm_{cluster_name}_{namespace}_{cm_name}"
                                            if not G.has_node(cm_id):
                                                make_node(cm_id, type="Environment", label=f"{cm_name}\\n(ConfigMap)", file_type="ConfigMap", source_file="-", shape='icon', icon={'face': '"Font Awesome 6 Free"', 'code': '\uf013', 'weight': '900', 'color': get_ns_color(namespace)}, provider="kubernetes", namespace=namespace, size=10)
                                            G.add_edge(pod_id, cm_id, relation="mounts_cm")
                                    if "secret" in v:
                                        sec_name = v["secret"].get("secretName")
                                        if sec_name:
                                            sec_id = f"sec_{cluster_name}_{namespace}_{sec_name}"
                                            if not G.has_node(sec_id):
                                                make_node(sec_id, type="Environment", label=f"{sec_name}\\n(Secret)", file_type="Secret", source_file="-", shape='icon', icon={'face': '"Font Awesome 6 Free"', 'code': '\uf084', 'weight': '900', 'color': get_ns_color(namespace)}, provider="kubernetes", namespace=namespace, size=10)
                                            G.add_edge(pod_id, sec_id, relation="mounts_secret")

                                for c in containers:
                                    env_from = c.get("envFrom", [])
                                    for ef in env_from:
                                        if "configMapRef" in ef:
                                            cm_name = ef["configMapRef"].get("name")
                                            if cm_name:
                                                cm_id = f"cm_{cluster_name}_{namespace}_{cm_name}"
                                                if not G.has_node(cm_id):
                                                    make_node(cm_id, type="Environment", label=f"{cm_name}\\n(ConfigMap)", file_type="ConfigMap", source_file="-", shape='icon', icon={'face': '"Font Awesome 6 Free"', 'code': '\uf013', 'weight': '900', 'color': get_ns_color(namespace)}, provider="kubernetes", namespace=namespace, size=10)
                                                G.add_edge(pod_id, cm_id, relation="uses_env_cm")
                                        if "secretRef" in ef:
                                            sec_name = ef["secretRef"].get("name")
                                            if sec_name:
                                                sec_id = f"sec_{cluster_name}_{namespace}_{sec_name}"
                                                if not G.has_node(sec_id):
                                                    make_node(sec_id, type="Environment", label=f"{sec_name}\\n(Secret)", file_type="Secret", source_file="-", shape='icon', icon={'face': '"Font Awesome 6 Free"', 'code': '\uf084', 'weight': '900', 'color': get_ns_color(namespace)}, provider="kubernetes", namespace=namespace, size=10)
                                                G.add_edge(pod_id, sec_id, relation="uses_env_secret")
                                                
                                    env = c.get("env", [])
                                    for ev in env:
                                        value_from = ev.get("valueFrom", {})
                                        if "configMapKeyRef" in value_from:
                                            cm_name = value_from["configMapKeyRef"].get("name")
                                            if cm_name:
                                                cm_id = f"cm_{cluster_name}_{namespace}_{cm_name}"
                                                if not G.has_node(cm_id):
                                                    make_node(cm_id, type="Environment", label=f"{cm_name}\\n(ConfigMap)", file_type="ConfigMap", source_file="-", shape='icon', icon={'face': '"Font Awesome 6 Free"', 'code': '\uf013', 'weight': '900', 'color': get_ns_color(namespace)}, provider="kubernetes", namespace=namespace, size=10)
                                                G.add_edge(pod_id, cm_id, relation="uses_env_cm")
                                        if "secretKeyRef" in value_from:
                                            sec_name = value_from["secretKeyRef"].get("name")
                                            if sec_name:
                                                sec_id = f"sec_{cluster_name}_{namespace}_{sec_name}"
                                                if not G.has_node(sec_id):
                                                    make_node(sec_id, type="Environment", label=f"{sec_name}\\n(Secret)", file_type="Secret", source_file="-", shape='icon', icon={'face': '"Font Awesome 6 Free"', 'code': '\uf084', 'weight': '900', 'color': get_ns_color(namespace)}, provider="kubernetes", namespace=namespace, size=10)
                                                G.add_edge(pod_id, sec_id, relation="uses_env_secret")
                            G.add_edge(ctrl_id, pod_id, relation="manages")

        # Fallback: Extract controllers directly from Pod ownerReferences (handles partial JSON dumps)
        for p in pods:
            pod_name = p["metadata"]["name"]
            namespace = p["metadata"].get("namespace", "default")
            pod_id = f"pod_{cluster_name}_{namespace}_{pod_name}"
            
            # Only process if this Pod was added to the graph
            if G.has_node(pod_id):
                sa_name = p.get("spec", {}).get("serviceAccountName")
                if sa_name:
                    sa_id = f"sa_{cluster_name}_{namespace}_{sa_name}"
                    if not G.has_node(sa_id):
                        make_node(sa_id, type="Environment", label=f"{sa_name}\\n(ServiceAccount)", file_type="ServiceAccount", source_file="-", shape='icon', icon={'face': '"Font Awesome 6 Free"', 'code': '\uf2c2', 'weight': '900', 'color': '#FF9F1C'}, provider="kubernetes", namespace=namespace, size=15)
                    G.add_edge(pod_id, sa_id, relation="uses_sa")

                node_name = p.get("spec", {}).get("nodeName")
                if node_name:
                    n_id = f"node_{cluster_name}_{node_name}"
                    if not G.has_node(n_id):
                        make_node(n_id, type="Server", label=f"{node_name}\\n(Server Node)", file_type="K8s Node", source_file="-", shape='icon', icon={'face': '"Font Awesome 6 Free"', 'code': '\uf233', 'weight': '900', 'color': '#118AB2'}, provider="kubernetes", namespace="system", size=25)
                    G.add_edge(pod_id, n_id, relation="scheduled_on")

                image_pull_secrets = p.get("spec", {}).get("imagePullSecrets", [])
                for ips in image_pull_secrets:
                    sec_name = ips.get("name")
                    if sec_name:
                        sec_id = f"sec_{cluster_name}_{namespace}_{sec_name}"
                        if not G.has_node(sec_id):
                            make_node(sec_id, type="Environment", label=f"{sec_name}\\n(Secret)", file_type="Secret", source_file="-", shape='icon', icon={'face': '"Font Awesome 6 Free"', 'code': '\uf084', 'weight': '900', 'color': get_ns_color(namespace)}, provider="kubernetes", namespace=namespace, size=10)
                        G.add_edge(pod_id, sec_id, relation="pulls_image_with_secret")

                owners = p["metadata"].get("ownerReferences", [])
                for owner in owners:
                    owner_kind = owner.get("kind")
                    owner_name = owner.get("name")
                    if not owner_kind or not owner_name: continue
                    
                    icon = '\uf0e8'
                    color = "#F15BB5"
                    if owner_kind == "StatefulSet":
                        icon = '\uf1c0'; color = "#FEE440"
                    elif owner_kind == "DaemonSet":
                        icon = '\uf233'; color = "#9B5DE5"
                    elif owner_kind == "Job":
                        icon = '\uf013'; color = "#3A86FF"
                    
                    ctrl_id = f"{owner_kind.lower()}_{cluster_name}_{namespace}_{owner_name}"
                    if not G.has_node(ctrl_id):
                        make_node(ctrl_id, type="Region", label=f"{owner_name}\\n({owner_kind})", file_type=f"K8s {owner_kind}", source_file=f"Namespace: {namespace}", shape='icon', icon={'face': '"Font Awesome 6 Free"', 'code': icon, 'weight': '900', 'color': color}, provider="kubernetes", namespace=namespace, size=18)
                    
                    G.add_edge(ctrl_id, pod_id, relation="manages")
                    
                    # Backlink so tracing downwards from a Service reaches the Controller
                    for src, tgt, edge_data in list(G.in_edges(pod_id, data=True)):
                        if edge_data.get("relation") == "selects_pod":
                            G.add_edge(src, ctrl_id, relation="targets_controller")

        # RBAC: Extract Roles and RoleBindings
        for rb in rolebindings + clusterrolebindings:
            rb_name = rb["metadata"]["name"]
            rb_ns = rb["metadata"].get("namespace", "default")
            
            role_ref = rb.get("roleRef", {})
            role_kind = role_ref.get("kind")
            role_name = role_ref.get("name")
            if not role_name: continue
            
            role_id = f"{role_kind.lower()}_{cluster_name}_{rb_ns}_{role_name}"
            if not G.has_node(role_id):
                make_node(role_id, type="Environment", label=f"{role_name}\\n({role_kind})", file_type=role_kind, source_file="-", shape='icon', icon={'face': '"Font Awesome 6 Free"', 'code': '\uf508', 'weight': '900', 'color': '#EF476F'}, provider="kubernetes", namespace=rb_ns, size=15)
            
            subjects = rb.get("subjects", [])
            if not subjects: continue
            for sub in subjects:
                if sub.get("kind") == "ServiceAccount":
                    sa_name = sub.get("name")
                    sa_ns = sub.get("namespace", rb_ns)
                    sa_id = f"sa_{cluster_name}_{sa_ns}_{sa_name}"
                    
                    if not G.has_node(sa_id):
                        make_node(sa_id, type="Environment", label=f"{sa_name}\\n(ServiceAccount)", file_type="ServiceAccount", source_file="-", shape='icon', icon={'face': '"Font Awesome 6 Free"', 'code': '\uf2c2', 'weight': '900', 'color': '#FF9F1C'}, provider="kubernetes", namespace=sa_ns, size=15)
                    
                    G.add_edge(sa_id, role_id, relation="bound_to")

        # Network Policies
        for np in networkpolicies:
            np_name = np["metadata"]["name"]
            np_ns = np["metadata"].get("namespace", "default")
            
            np_id = f"np_{cluster_name}_{np_ns}_{np_name}"
            if not G.has_node(np_id):
                make_node(np_id, type="Environment", label=f"{np_name}\\n(NetworkPolicy)", file_type="NetworkPolicy", source_file="-", shape='icon', icon={'face': '"Font Awesome 6 Free"', 'code': '\uf3ed', 'weight': '900', 'color': '#06D6A0'}, provider="kubernetes", namespace=np_ns, size=18)
            
            pod_selector = np.get("spec", {}).get("podSelector", {}).get("matchLabels", {})
            for p in pods:
                if p["metadata"].get("namespace", "default") != np_ns:
                    continue
                labels = p["metadata"].get("labels", {})
                match = True
                for k, v in pod_selector.items():
                    if labels.get(k) != v:
                        match = False
                        break
                if match:
                    pod_name = p["metadata"]["name"]
                    pod_id = f"pod_{cluster_name}_{np_ns}_{pod_name}"
                    if G.has_node(pod_id):
                        G.add_edge(np_id, pod_id, relation="secures_pod", dashes=True, color="#06D6A0", width=2)
            
            ingress = np.get("spec", {}).get("ingress", [])
            for rule in ingress:
                for fr in rule.get("from", []):
                    pod_sel = fr.get("podSelector", {}).get("matchLabels", {})
                    if pod_sel:
                        for p in pods:
                            if p["metadata"].get("namespace", "default") != np_ns:
                                continue
                            labels = p["metadata"].get("labels", {})
                            match = True
                            for k, v in pod_sel.items():
                                if labels.get(k) != v:
                                    match = False
                                    break
                            if match:
                                pod_name = p["metadata"]["name"]
                                pod_id = f"pod_{cluster_name}_{np_ns}_{pod_name}"
                                if G.has_node(pod_id):
                                    G.add_edge(pod_id, np_id, relation="allows_ingress", dashes=True, color="#06D6A0", width=1)
            
            egress = np.get("spec", {}).get("egress", [])
            for rule in egress:
                for to in rule.get("to", []):
                    pod_sel = to.get("podSelector", {}).get("matchLabels", {})
                    if pod_sel:
                        for p in pods:
                            if p["metadata"].get("namespace", "default") != np_ns:
                                continue
                            labels = p["metadata"].get("labels", {})
                            match = True
                            for k, v in pod_sel.items():
                                if labels.get(k) != v:
                                    match = False
                                    break
                            if match:
                                pod_name = p["metadata"]["name"]
                                pod_id = f"pod_{cluster_name}_{np_ns}_{pod_name}"
                                if G.has_node(pod_id):
                                    G.add_edge(np_id, pod_id, relation="allows_egress", dashes=True, color="#EF476F", width=1)

def _compute_communities(G):
    type_to_cid = {
        "Server": 0,
        "Region": 1,
        "Environment": 2,
        "MainDomain": 3,
        "SubDomain": 4,
    }
    
    community_labels = {cid: label for label, cid in type_to_cid.items()}
    communities = defaultdict(list)
    namespace_to_cid = {}
    
    for node_id, data in G.nodes(data=True):
        ns = data.get("namespace")
        if ns:
            if ns not in namespace_to_cid:
                cid = len(type_to_cid) + len(namespace_to_cid) + 100
                namespace_to_cid[ns] = cid
                community_labels[cid] = f"NS: {ns}"
            cid = namespace_to_cid[ns]
        else:
            ntype = data.get("type", "Unknown")
            if ntype not in type_to_cid:
                cid = len(type_to_cid)
                type_to_cid[ntype] = cid
                community_labels[cid] = ntype
            else:
                cid = type_to_cid[ntype]
        communities[cid].append(node_id)
        
    G.graph["hyperedges"] = []
    return dict(communities), community_labels


def build_devops_graph(data_dir: str, k8s_only: bool = False):
    graphs = {}
    
    k8s_dir = os.path.join(data_dir, "kubernetes")
    cluster_files = []
    if os.path.exists(k8s_dir):
        cluster_files = sorted([f for f in os.listdir(k8s_dir) if f.endswith(".json")])
        
    if cluster_files:
        for filename in cluster_files:
            cluster_name = filename.replace('.json', '')
            graph_name = f"graph_{cluster_name}"
            
            G = nx.DiGraph()
            servers_by_ip = {}
            if not k8s_only:
                parse_linode(data_dir, G, servers_by_ip)
                parse_bunny(data_dir, G, servers_by_ip)
            
            parse_kubernetes(data_dir, G, servers_by_ip, cluster_filename=filename)
            
            communities, community_labels = _compute_communities(G)
            graphs[graph_name] = (G, communities, community_labels)
    else:
        G = nx.DiGraph()
        servers_by_ip = {}
        if not k8s_only:
            parse_linode(data_dir, G, servers_by_ip)
            parse_bunny(data_dir, G, servers_by_ip)
            
        communities, community_labels = _compute_communities(G)
        graphs["graph"] = (G, communities, community_labels)
        
    return graphs
