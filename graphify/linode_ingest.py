import os
import csv
import glob

def parse_linode(data_dir, G, servers_by_ip):
    linode_dir = os.path.join(data_dir, 'linode')
    if not os.path.exists(linode_dir):
        return
        
    G.add_node("Linode Cloud", type="Cloud", label="Linode Cloud\n(Provider)", file_type="Cloud", source_file="-",
               shape='icon', icon={'face': '"Font Awesome 6 Free"', 'code': '\uf0c2', 'weight': '900', 'color': '#00A95C'}, provider="linode")
        
    # Support CSV files
    linode_files = glob.glob(os.path.join(linode_dir, '*.csv'))
    for linodes_file in linode_files:
        with open(linodes_file, 'r', encoding='utf-8') as f:
            sample = f.read(1024)
            f.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample)
                reader = csv.DictReader(f, dialect=dialect)
            except csv.Error:
                # Fallback to standard comma separated if Sniffer can't determine it
                reader = csv.DictReader(f)
                
            for row in reader:
                # Normalize headers to lowercase to match Excel parser behavior and prevent bugs
                row_dict = {str(k).strip().lower() if k else f"col_{i}": str(v) for i, (k, v) in enumerate(row.items())}
                _process_linode_row(row_dict, G, servers_by_ip)
                
    # Support Excel files (.xlsx)
    xlsx_files = glob.glob(os.path.join(linode_dir, '*.xlsx'))
    if xlsx_files:
        try:
            import openpyxl
            for xlsx_file in xlsx_files:
                wb = openpyxl.load_workbook(xlsx_file, data_only=True)
                ws = wb.active
                headers = [str(cell.value).strip().lower() if cell.value else f"col_{i}" for i, cell in enumerate(ws[1])]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_dict = dict(zip(headers, [str(v) if v is not None else "" for v in row]))
                    _process_linode_row(row_dict, G, servers_by_ip)
        except ImportError:
            print("Warning: openpyxl is not installed. Please run `pip install openpyxl` to parse Linode.xlsx.")

def _process_linode_row(row, G, servers_by_ip):
    if not row or 'label' not in row or not row['label'].strip():
        return
        
    name = row['label']
    ips_raw = row.get('ipv4', '')
    region = row.get('region', '')
    tag = row.get('tags', '')
    
    image = row.get('image', '')
    status = row.get('status', '')
    instance_type = row.get('type', '')
    vcpus = row.get('vcpus', '')
    
    ips = [ip.strip(' "') for ip in ips_raw.split(',') if ip.strip(' "')]
    ip_str = ", ".join(ips)
    
    source_info = (
        f"IP(s): {ip_str} | Region: {region} | Tags: {tag} | "
        f"Image: {image} | Status: {status} | InstType: {instance_type} | vCPUs: {vcpus}"
    )
    
    node_label = f"{name}\n({ip_str})\n[{instance_type}]" if instance_type else f"{name}\n({ip_str})"
    
    status_lower = status.lower() if status else ""
    if "offline" in status_lower or "stopped" in status_lower:
        icon_color = "#06D6A0"  # Green for down
    elif "running" in status_lower:
        icon_color = "#4E79A7"  # Blue for running
    elif status_lower:
        icon_color = "#FFCA3A"  # Yellow for booting/provisioning
    else:
        icon_color = "#4E79A7"  # Default blue if no status

    G.add_node(name, 
               type="Server", 
               ip=ip_str,
               label=node_label,
               color=icon_color,
               file_type="Linode Server",
               source_file=source_info,
               shape='icon',
               icon={'face': '"Font Awesome 6 Free"', 'code': '\uf233', 'weight': '900', 'color': icon_color},
               provider="linode")
    
    for ip in ips:
        servers_by_ip[ip] = name
    
    if region:
        G.add_node(region, type="Region", label=f"{region}\n(Region)", file_type="Region", source_file="-",
                   shape='icon', icon={'face': '"Font Awesome 6 Free"', 'code': '\uf3c5', 'weight': '900', 'color': '#F28E2B'}, provider="linode")
        G.add_edge("Linode Cloud", region, relation="has_region", color="#00A95C")

    if tag:
        # Create unique tag ID per region to ensure a perfect sunburst tree instead of crossing lines
        tag_id = f"{region}_{tag}" if region else tag
        G.add_node(tag_id, type="Environment", label=f"{tag}\n(Environment)", file_type="Environment", source_file="-",
                   shape='icon', icon={'face': '"Font Awesome 6 Free"', 'code': '\uf02b', 'weight': '900', 'color': '#E15759'}, provider="linode")
                   
        # Tag attaches to region if it exists, otherwise directly to Linode Cloud
        if region:
            G.add_edge(region, tag_id, relation="has_tag", color="#F28E2B")
        else:
            G.add_edge("Linode Cloud", tag_id, relation="has_tag", color="#00A95C")
            
        # Server attaches to tag
        G.add_edge(tag_id, name, relation="contains_server", color="#E15759")
    else:
        # If no tag, Server attaches to region, or directly to Linode Cloud
        if region:
            G.add_edge(region, name, relation="deployed_in", color="#F28E2B")
        else:
            G.add_edge("Linode Cloud", name, relation="contains_server", color="#00A95C")

